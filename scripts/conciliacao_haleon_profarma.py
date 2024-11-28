import pandas as pd
import openpyxl
import sys

# Função para remover as aspas simples do CNPJ
def remover_aspas_cnpj(cnpj):
    # Remover aspas simples do CNPJ
    return str(cnpj).replace("'", "")

# Funções de formatação e remoção de sufixo decimal
def format_cnpj(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(value) if isinstance(value, str) else str(value)
    return value.zfill(14)

def remove_decimal_suffix(value):
    if isinstance(value, str) and value.endswith('.0'):
        return value[:-2]
    return value

# Definição dos arquivos de origem e destino
arquivo_distribuidor = sys.argv[2] 
arquivo_funcional = sys.argv[1] 
arquivo_valor_maximo = sys.argv[3]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)

# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)

# 3. Carregar o arquivo arquivo valor maximo
print('Lendo o arquivo valor maximo...')
try:
    valor_maximo = pd.read_excel(arquivo_valor_maximo)
except Exception as e:
    print(f"Erro ao carregar o arquivo valor máximo: {e}")
    sys.exit(1)

# Criar novas colunas para o status de validação, nova chave, e valor máximo de ressarcimento
base_distribuidor['Status Validação'] = ''
base_distribuidor['CNPJ+NF+Novo EAN'] = ''
base_distribuidor['Valor máximo ressarcimento'] = ''

# Dicionário DE-PARA
de_para = {
    '7896009498961': '7891045043588',
    '7896015590932': '7896015516024',
    '7896015516048': '7896015590956',
    '7896015570255': '7896015590987',
    '7891045043588': '7896009498398',
    '7896009498732': '7896015529222'
}

# Criar chaves e filtros para base_funcional
base_funcional['Chave EAN+NF+CNPJ'] = base_funcional[['EAN do produto', 'Nota Fiscal', 'CNPJ']].astype(str).agg(''.join, axis=1)
base_funcional['Chave NF+CNPJ'] = base_funcional[['Nota Fiscal, CNPJ']].astype(str).agg(''.join, axis=1)
base_funcional['Chave NF'] = base_funcional['Nota Fiscal']

# Criar chave EAN+Estado para valor_maximo
valor_maximo['EAN+UF'] = valor_maximo[['EAN + UF']].astype(str).agg(''.join, axis=1)

# Verificar presença e atribuição de status
base_distribuidor['Chave EAN+NF+CNPJ'] = base_distribuidor[['EAN', 'NF', 'CNPJ']].astype(str).agg(''.join, axis=1)
base_distribuidor['Chave NF'] = base_distribuidor['NF']

# Encontrar EAN divergente e substituição
base_distribuidor['Status Validação'] = base_distribuidor['Chave EAN+NF+CNPJ'].map(
    lambda chave: 'Encontrado' if chave in base_funcional['Chave EAN+NF+CNPJ'].values else
    ('EAN divergente' if base_distribuidor['Chave NF'].values[0] in base_funcional['Chave NF'].values else 'NF não Trafegou')
)

# Substituir EAN se necessário
base_distribuidor.loc[base_distribuidor['Status Validação'] == 'EAN divergente', 'CNPJ+NF+Novo EAN'] = base_distribuidor['EAN'].map(de_para).fillna('')

# Verificar presença de novo EAN
base_distribuidor['Status Validação'] = base_distribuidor.apply(
    lambda row: 'Encontrado com Novo EAN' if row['CNPJ+NF+Novo EAN'] in base_funcional['Chave EAN+NF+CNPJ'].values else
    'Novo EAN não encontrado', axis=1
)

# Atribuir valor máximo de ressarcimento
base_distribuidor.loc[base_distribuidor['Unidades de Negócios '] == 'OTC', 'Valor máximo ressarcimento'] = base_distribuidor.apply(
    lambda row: base_funcional[(base_funcional['Nota Fiscal'].astype(str) == str(row['NF'])) & 
                              (base_funcional['CNPJ'].astype(str) == str(row['CNPJ']))]['Ressarcimento'].max(), axis=1
)

# Atribuir preço máximo Junho 2024
base_distribuidor.loc[base_distribuidor['Unidades de Negócios '] != 'OTC', 'Valor máximo ressarcimento'] = base_distribuidor.apply(
    lambda row: valor_maximo[valor_maximo['EAN+UF'] == str(row['EAN']) + str(row['Estado'])]['Preço máximo Junho 2024'].max(), axis=1
)

# salvar em duas abas diferentes do mesmo arquivo
with pd.ExcelWriter('Conciliacao_Finalizada_Santa_Cruz.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print('Finalizado!')