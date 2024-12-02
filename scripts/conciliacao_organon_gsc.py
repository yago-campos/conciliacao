import pandas as pd
import openpyxl
from collections import defaultdict
import sys

# Função para remover as aspas simples do CNPJ
def remover_aspas_cnpj(cnpj):
    return str(cnpj).replace("'", "")

# Funções de formatação e remoção de sufixo decimal
def format_cnpj(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(value) if isinstance(value, str) else str(value)
    return value.zfill(14)

# Função para remover decimal
def remove_decimal_suffix(value):
    if isinstance(value, float):
        return str(int(value))  # Converte para int e depois para string
    return str(value)

def mapear_chave_secundaria(row, base):
    # Identifica as colunas de acordo com a base
    if 'Nr. Nota' in row:  # base_distribuidor
        nf = str(row['Nr. Nota']).strip()  # Convertendo para string e removendo espaços
        cnpj_nf = str(row['CNPJ']) + '-' + nf
        cnpj_nf_ean = cnpj_nf + '-' + str(row['Código EAN']).strip()
        qtd_distribuidor = row['Quantidade Somada']
    else:  # base_funcional
        nf = str(row['Número da nota']).strip()
        cnpj_nf = str(row['CNPJ']) + '-' + nf
        cnpj_nf_ean = cnpj_nf + '-' + str(row['EAN do produto']).strip()
        qtd_distribuidor = None  # Quantidade não relevante aqui

    # 1. Verifica se o número da nota existe
    if nf in base['Número da nota'].astype(str).values:
        # Encontra a nota: continua com a verificação
        status = "NF ok, verificando outros campos"
        
        # 2. Verifica CNPJ+NF
        if cnpj_nf in base['CNPJ+NF'].values:
            # CNPJ+NF encontrados: continua com a verificação
            status = "CNPJ+NF ok, verificando outros campos"
            
            # 3. Verifica CNPJ+NF+EAN (sem considerar a quantidade ainda)
            if cnpj_nf_ean in base['CNPJ+NF+EAN'].values:
                # Encontra EAN: agora verifica se a quantidade diverge
                qtd_funcional = base.loc[base['CNPJ+NF+EAN'] == cnpj_nf_ean, 'Quantidade Somada'].values[0]
                
                # Verifica se a quantidade é diferente
                if qtd_distribuidor is not None and qtd_distribuidor != qtd_funcional:
                    return "CNPJ+NF+EAN ok, qtd divergente"
                else:
                    return "CNPJ+NF+EAN ok"
            else:
                return "EAN ou Quantidade divergente"  # EAN não encontrado
        else:
            return "Produto fora da NF"  # CNPJ+NF não encontrado
    else:
        return "NF não localizada"  # Nota não encontrada

# Definição dos arquivos de origem e destino
arquivo_funcional = sys.argv[1]
arquivo_distribuidor = sys.argv[2]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)


# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)

# Formatar CNPJ nas bases
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(remover_aspas_cnpj).apply(format_cnpj).astype(str)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(remove_decimal_suffix).apply(format_cnpj).astype(str)

# Criar a quantidade somada antes de definir a chave para base funcional
base_funcional['Quantidade Somada'] = base_funcional.groupby(['CNPJ', 'Número da nota', 'EAN do produto'])['Quantidade Faturada'].transform('sum')

# Criar a chave com a quantidade somada para base funcional
base_funcional['Chave'] = (
    base_funcional['CNPJ'] + '-' +
    base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_funcional['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
).astype(str)  # Assegura que a chave seja uma string

# Criar a quantidade somada antes de definir a chave para base distribuidor
base_distribuidor['Quantidade Somada'] = base_distribuidor.groupby(['CNPJ', 'Nr. Nota', 'Código EAN'])['Qtd'].transform('sum')

# Remover .0 da chave na base funcional
base_funcional['Chave'] = base_funcional['Chave'].str.replace('.0', '', regex=False)

# Criar a chave com a quantidade somada para base distribuidor
base_distribuidor['Chave'] = (
    base_distribuidor['CNPJ'] + '-' +
    base_distribuidor['Nr. Nota'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_distribuidor['Código EAN'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_distribuidor['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
).astype(str)  # Assegura que a chave seja uma string

# Remover .0 da chave na base funcional
base_distribuidor['Chave'] = base_distribuidor['Chave'].str.replace('.0', '', regex=False)

# Criar as colunas CNPJ+NF e CNPJ+NF+EAN
base_distribuidor['CNPJ+NF'] = base_distribuidor['CNPJ'].astype(str) + '-' + base_distribuidor['Nr. Nota'].astype(str)
base_funcional['CNPJ+NF'] = base_funcional['CNPJ'].astype(str) + '-' + base_funcional['Número da nota'].astype(str)
base_funcional['CNPJ+NF+EAN'] = base_funcional['Número da nota'].astype(str) + '-' + base_funcional['EAN do produto'].astype(str)

# Comparando as bases
print('Comparando as bases...')
base_distribuidor['Status'] = base_distribuidor['Chave'].isin(base_funcional['Chave']).map({True: 'OK', False: 'Chave não localizada'})
base_funcional['Status'] = base_funcional['Chave'].isin(base_distribuidor['Chave']).map({True: 'OK', False: 'Chave só consta na Funcional'})

# Aplicação do mapeamento adicional quando "Chave não localizada"
base_distribuidor['Status'] = base_distribuidor.apply(
    lambda row: mapear_chave_secundaria(row, base_funcional) if row['Status'] == 'Chave não localizada' else row['Status'],
    axis=1
)

# "Chave só consta na Funcional"
base_funcional['Status'] = base_funcional.apply(
    lambda row: 'Chave só consta no Distribuidor' if row['Status'] == 'Chave não localizada' else row['Status'],
    axis=1
)

# Reordenando colunas
novas_colunas_funcional = [
    'Chave','Status'
]
outras_colunas_funcional = [col for col in base_funcional.columns if col not in novas_colunas_funcional]
base_funcional = base_funcional[novas_colunas_funcional + outras_colunas_funcional]

novas_colunas_distribuidor = [
    'Chave', 'Status'
]
outras_colunas_distribuidor = [col for col in base_distribuidor.columns if col not in novas_colunas_distribuidor]
base_distribuidor = base_distribuidor[novas_colunas_distribuidor + outras_colunas_distribuidor]

# 1. Criar o dicionário 'dados_funcional' com a chave e as colunas necessárias
dados_funcional = {}

# Preencher o dicionário com as informações da base_funcional
for _, row in base_funcional[['Chave', 'Desconto pedido', 'Preço Fábrica']].iterrows():
    if row['Chave'] not in dados_funcional:
        dados_funcional[row['Chave']] = []
    dados_funcional[row['Chave']].append({
        'Desconto pedido': row['Desconto pedido'],
        'Preço Fábrica': row['Preço Fábrica']
    })

# Calcular o 'Ressarcimento Funcional' usando 'Qtd' da base_distribuidor
def calcular_ressarcimento(row):
    chave = row['Chave']
    qtd_distribuidor = row['Qtd']  # Quantidade individual da base_distribuidor
    
    # Verificar se a chave existe nos dados_funcional
    if chave in dados_funcional:
        # Pegar o primeiro item correspondente à chave
        item = dados_funcional[chave][0]  # Seleciona o primeiro item da lista
        # Calcular o ressarcimento usando esse único item
        total_ressarcimento = item['Desconto pedido'] * item['Preço Fábrica'] * qtd_distribuidor / 100
        return total_ressarcimento
    return 0  # Se não encontrar a chave, retorna 0

# Aplicar a função na base_distribuidor
base_distribuidor['Ressarcimento Funcional'] = base_distribuidor.apply(calcular_ressarcimento, axis=1)

base_funcional.drop(columns=['CNPJ+NF', 'CNPJ+NF+EAN', 'Quantidade Somada'], inplace=True)
base_distribuidor.drop(columns=['CNPJ+NF', 'Quantidade Somada'], inplace=True)

# salvar em duas abas diferentes do mesmo arquivo
with pd.ExcelWriter('Conciliacao_Finalizada_Santa_Cruz.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print("Processo concluído com sucesso!")