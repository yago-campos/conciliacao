import pandas as pd
import openpyxl
import os
import time
import sys

# Função para formatar o CNPJ para 14 dígitos
def format_cnpj(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(int(value)) if isinstance(value, (int, float)) else str(value)
    return value.zfill(14)

# Função para remover o sufixo ".0"
def remove_decimal_suffix(value):
    if isinstance(value, str) and value.endswith('.0'):
        return value[:-2]
    return value

# Função para remover o sufixo após a barra "/"
def remove_suffix_after_slash(value):
    if isinstance(value, str) and '/' in value:
        value = value.split('/')[0].strip()  # Remove o sufixo após a barra e remove espaços extras
        try:
            return int(value)  # Tenta converter para número inteiro
        except ValueError:
            return value  # Retorna o valor original se não for possível converter
    return value

# Função para atualizar o status e remover colunas desnecessárias
def atualizar_status_remover_colunas(df):
    df.loc[df['Validação V/F'] == True, 'Status'] = df['Status1']
    df = df.drop(columns=['Status1', 'Status2', 'Validação V/F'])
    return df

# Definição dos arquivos de origem e destino
arquivo_distribuidor = sys.argv[2]
arquivo_funcional = sys.argv[1]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)

# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)
    
# Formatar CNPJ na base Funcional e Distribuidor para 14 dígitos
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(format_cnpj)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(format_cnpj)
base_distribuidor['Nota Fiscal'] = base_distribuidor['Nota Fiscal'].apply(format_cnpj).apply(remove_suffix_after_slash)

# Criar colunas específicas combinadas nas bases
print('Criando colunas combinadas nas bases Funcional e Distribuidor...')

base_funcional['CNPJ+NF'] = base_funcional['CNPJ'] + '-' + base_funcional['Nota Fiscal'].astype(str).apply(remove_decimal_suffix)
base_funcional['CNPJ+NF+EAN'] = base_funcional['CNPJ+NF'] + '-' + base_funcional['EAN Produto'].astype(str).apply(remove_decimal_suffix)

base_distribuidor['CNPJ+NF'] = base_distribuidor['CNPJ'] + '-' + base_distribuidor['Nota Fiscal'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['CNPJ+NF+EAN'] = base_distribuidor['CNPJ+NF'] + '-' + base_distribuidor['EAN Produto'].astype(str).apply(remove_decimal_suffix)

# Criar chaves na base Funcional
print('Criando chaves na base Funcional...')
base_funcional['Chave1'] = (
base_funcional['CNPJ'] + '-' +
base_funcional['Nota Fiscal'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['EAN Produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['Quantidade Faturada'].astype(str).apply(remove_decimal_suffix)
)
base_funcional['Quantidade Somada'] = base_funcional.groupby(['CNPJ', 'Nota Fiscal', 'EAN Produto'])['Quantidade Faturada'].transform('sum')
base_funcional['Chave2'] = (
base_funcional['CNPJ'] + '-' +
base_funcional['Nota Fiscal'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['EAN Produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)
base_funcional['Status'] = ''
base_funcional['Status1'] = ''
base_funcional['Status2'] = ''
base_funcional['Validação V/F'] = ''
base_funcional['AÇÃO'] = ''
base_funcional['ORIGEM'] = ''

base_funcional['ORIGEM'] = base_funcional['Origem do pedido'].apply(lambda x: 'CA' if x == 'Canal Autorizador' else 'WE')

# Criar chaves na base do Distribuidor
print('Criando chaves na base do Distribuidor...')
base_distribuidor['Chave1'] = (
base_distribuidor['CNPJ'] + '-' +
base_distribuidor['Nota Fiscal'].astype(str).apply(remove_decimal_suffix).apply(remove_suffix_after_slash) + '-' +
base_distribuidor['EAN Produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['Qtd. Vendida'].astype(str).apply(remove_decimal_suffix)
)
base_distribuidor['Quantidade Somada'] = base_distribuidor.groupby(['CNPJ', 'Nota Fiscal', 'EAN Produto'])['Qtd. Vendida'].transform('sum')
base_distribuidor['Chave2'] = (
base_distribuidor['CNPJ'] + '-' +
base_distribuidor['Nota Fiscal'].astype(str).apply(remove_decimal_suffix).apply(remove_suffix_after_slash) + '-' +
base_distribuidor['EAN Produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)
base_distribuidor['Status'] = ''
base_distribuidor['Status1'] = ''
base_distribuidor['Status2'] = ''
base_distribuidor['Validação V/F'] = ''
base_distribuidor['AÇÃO'] = ''
base_distribuidor['ORIGEM'] = ''

# Comparação das bases
print('Comparando as bases Funcional e Distribuidor...')
base_distribuidor['Status1'] = base_distribuidor['Chave1'].isin(base_funcional['Chave1']).map({True: 'NF ok', False: 'NF não localizada'})
base_funcional['Status1'] = base_funcional['Chave1'].isin(base_distribuidor['Chave1']).map({True: 'NF ok', False: 'NF só consta na Funcional'})

base_distribuidor['Status2'] = base_distribuidor['Chave2'].isin(base_funcional['Chave2']).map({True: 'NF ok', False: 'NF não localizada'})
base_funcional['Status2'] = base_funcional['Chave2'].isin(base_distribuidor['Chave2']).map({True: 'NF ok', False: 'NF só consta na Funcional'})

# Validação V/F
base_funcional['Validação V/F'] = base_funcional['Status1'].eq(base_funcional['Status2']).map({True: 'OK', False: 'Falso'})
base_distribuidor['Validação V/F'] = base_distribuidor['Status1'].eq(base_distribuidor['Status2']).map({True: 'OK', False: 'Falso'})

# Atualizar status e remover colunas desnecessárias
base_funcional = atualizar_status_remover_colunas(base_funcional)
base_distribuidor = atualizar_status_remover_colunas(base_distribuidor)

# Criar coluna "Procx da NF"
def procx_da_nota(nota):
    match = base_funcional.loc[base_funcional['Nota Fiscal'] == nota, 'Nota Fiscal']
    if not match.empty:
        return match.iloc[0]
    return 'Nota Fiscal não localizada na Funcional'

base_distribuidor['Procx da NF'] = base_distribuidor.apply(
lambda row: procx_da_nota(row['Nota Fiscal']) if row['Status'] == 'NF não localizada' else '',
axis=1
)

# Criar colunas PROCV para CNPJ+NF e CNPJ+NF+EAN
base_distribuidor['PROCV CNPJ+NF'] = base_distribuidor['CNPJ+NF'].isin(base_funcional['CNPJ+NF']).map({True: 'Encontrado', False: 'Não encontrado'})
base_distribuidor['PROCV CNPJ+NF+EAN'] = base_distribuidor['CNPJ+NF+EAN'].isin(base_funcional['CNPJ+NF+EAN']).map({True: 'Encontrado', False: 'Não encontrado'})

# Reordenar colunas
novas_colunas_funcional = [
'Chave1', 'Chave2', 'Quantidade Somada', 'CNPJ+NF', 'CNPJ+NF+EAN', 'Status', 'AÇÃO', 'ORIGEM'
]
outras_colunas_funcional = [col for col in base_funcional.columns if col not in novas_colunas_funcional]
base_funcional = base_funcional[novas_colunas_funcional + outras_colunas_funcional]

novas_colunas_distribuidor = [
'Chave1', 'Chave2', 'Quantidade Somada', 'CNPJ+NF', 'CNPJ+NF+EAN', 'Status', 'Procx da NF', 'PROCV CNPJ+NF', 'PROCV CNPJ+NF+EAN'
]
outras_colunas_distribuidor = [col for col in base_distribuidor.columns if col not in novas_colunas_distribuidor]
base_distribuidor = base_distribuidor[novas_colunas_distribuidor + outras_colunas_distribuidor]

# salvar em duas abas diferentes do mesmo arquivo
with pd.ExcelWriter('Conciliacao_Finalizada_Santa_Cruz.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print('Finalizado!')