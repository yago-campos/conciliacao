import pandas as pd
import os
import openpyxl
import time
import sys

# Função para remover as aspas simples do CNPJ
def remover_aspas_cnpj(cnpj):
    # Remover aspas simples do CNPJ
    return str(cnpj).replace("'", "")

# Funções de formatação e remoção de sufixo decimal
def format_cnpj(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(value) if isinstance(value, str) else str(value)
    return value.zfill(14)

def remove_decimal_suffix(value):
    if isinstance(value, str) and value.endswith('.0'):
        return value[:-2]
    return value

# Função para atualizar o status e remover colunas desnecessárias
def atualizar_status_remover_colunas(df):
    df.loc[df['Validação V/F'] == True, 'Status'] = df['Status1']
    df = df.drop(columns=['Status1', 'Status2', 'Validação V/F'])
    return df

# Definição dos arquivos de origem e destino
arquivo_distribuidor = sys.argv[2]
arquivo_funcional = sys.argv[1]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)

# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)

# Formatar CNPJ nas bases
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(format_cnpj)
base_distribuidor['CNPJ_CLIENTE'] = base_distribuidor['CNPJ_CLIENTE'].apply(format_cnpj)

# Criar colunas combinadas nas bases
print('Criando as colunas combinadas nas bases...')
base_funcional['CNPJ+NF'] = base_funcional['CNPJ'] + '-' + base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix)
base_funcional['CNPJ+NF+EAN'] = base_funcional['CNPJ+NF'] + '-' + base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['CNPJ+NF'] = base_distribuidor['CNPJ_CLIENTE'] + '-' + base_distribuidor['NOTA'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['CNPJ+NF+EAN'] = base_distribuidor['CNPJ+NF'] + '-' + base_distribuidor['COD_EAN'].astype(str).apply(remove_decimal_suffix)

# Criar as chaves e colunas adicionais nas bases
print('Criando as chaves nas bases...')
base_funcional['Chave1'] = (
base_funcional['CNPJ'] + '-' +
base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['Quantidade Faturada'].astype(str).apply(remove_decimal_suffix)
)
base_funcional['Quantidade Somada'] = base_funcional.groupby(['CNPJ', 'Número da nota', 'EAN do produto'])['Quantidade Faturada'].transform('sum')
base_funcional['Chave2'] = (
base_funcional['CNPJ'] + '-' +
base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)
base_funcional['Status'] = ''
base_funcional['Status1'] = '' # excluir no final
base_funcional['Status2'] = '' # excluir no final
base_funcional['Validação V/F'] = '' # excluir no final
base_funcional['AÇÃO'] = ''
base_funcional['ORIGEM'] = ''

base_distribuidor['Chave1'] = (
base_distribuidor['CNPJ_CLIENTE'] + '-' +
base_distribuidor['NOTA'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['COD_EAN'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['QTD'].astype(str).apply(remove_decimal_suffix)
)
base_distribuidor['Quantidade Somada'] = base_distribuidor.groupby(['CNPJ_CLIENTE', 'NOTA', 'COD_EAN'])['QTD'].transform('sum')
base_distribuidor['Chave2'] = (
base_distribuidor['CNPJ_CLIENTE'] + '-' +
base_distribuidor['NOTA'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['COD_EAN'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)
base_distribuidor['Status'] = ''
base_distribuidor['Status1'] = '' # excluir no final
base_distribuidor['Status2'] = '' # excluir no final
base_distribuidor['Validação V/F'] = '' # excluir no final
base_distribuidor['AÇÃO'] = ''
base_distribuidor['ORIGEM'] = ''

# Comparando as bases
print('Comparando as bases...')
base_distribuidor['Status1'] = base_distribuidor['Chave1'].isin(base_funcional['Chave1']).map({True: 'OK', False: 'NF não localizada'})
base_funcional['Status1'] = base_funcional['Chave1'].isin(base_distribuidor['Chave1']).map({True: 'OK', False: 'NF só consta na Funcional'})

base_distribuidor['Status2'] = base_distribuidor['Chave2'].isin(base_funcional['Chave2']).map({True: 'OK', False: 'NF não localizada'})
base_funcional['Status2'] = base_funcional['Chave2'].isin(base_distribuidor['Chave2']).map({True: 'OK', False: 'NF só consta na Funcional'})

base_funcional['Validação V/F'] = base_funcional['Status1'].eq(base_funcional['Status2'])
base_distribuidor['Validação V/F'] = base_distribuidor['Status1'].eq(base_distribuidor['Status2'])

# Atualizando o status e removendo colunas desnecessárias
base_funcional = atualizar_status_remover_colunas(base_funcional)
base_distribuidor = atualizar_status_remover_colunas(base_distribuidor)

# Criar coluna de validação "Procx da NF"
def procx_da_nota(nota):
    match = base_funcional.loc[base_funcional['Número da nota'] == nota, 'Número da nota']
    if not match.empty:
        return match.iloc[0]
    return 'Número da nota também não localizado na Funcional'

base_distribuidor['Procx da NF'] = base_distribuidor.apply(
lambda row: procx_da_nota(row['NOTA']) if row['Status'] == 'NF não localizada' else '',
axis=1
)

# Criar colunas PROCV para CNPJ+NF e CNPJ+NF+EAN
base_distribuidor['PROCV CNPJ+NF'] = base_distribuidor['CNPJ+NF'].isin(base_funcional['CNPJ+NF']).map({True: 'Encontrado', False: 'Não encontrado'})
base_distribuidor['PROCV CNPJ+NF+EAN'] = base_distribuidor['CNPJ+NF+EAN'].isin(base_funcional['CNPJ+NF+EAN']).map({True: 'Encontrado', False: 'Não encontrado'})

# Reordenando colunas - Colocando as novas colunas no início
novas_colunas_funcional = [
'CNPJ+NF', 'CNPJ+NF+EAN', 'Chave1', 'Chave2', 'Quantidade Somada', 'Status', 'AÇÃO', 'ORIGEM'
]
outras_colunas_funcional = [col for col in base_funcional.columns if col not in novas_colunas_funcional]
base_funcional = base_funcional[novas_colunas_funcional + outras_colunas_funcional]

novas_colunas_distribuidor = [
'CNPJ+NF', 'CNPJ+NF+EAN', 'Chave1', 'Chave2', 'Quantidade Somada', 'Status', 'Procx da NF', 'PROCV CNPJ+NF', 'PROCV CNPJ+NF+EAN'
]
outras_colunas_distribuidor = [col for col in base_distribuidor.columns if col not in novas_colunas_distribuidor]
base_distribuidor = base_distribuidor[novas_colunas_distribuidor + outras_colunas_distribuidor]

# salvar em duas abas diferentes do mesmo arquivo
with pd.ExcelWriter('Conciliacao_Finalizada_Santa_Cruz.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print('Finalizado!')