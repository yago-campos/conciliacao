import pandas as pd
import os
import openpyxl
import time
import numpy as np
import re
import sys

# Função para remover as aspas simples do CNPJ
def remover_aspas_cnpj(cnpj):
    # Remover aspas simples do CNPJ
    return re.sub(r'[./-]', '', str(cnpj))

# Funções de formatação e remoção de sufixo decimal
def format_cnpj(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(value) if isinstance(value, str) else str(value)
    return value.zfill(14)

def remove_decimal_suffix(value):
    if isinstance(value, str) and value.endswith('.0'):
        return value[:-2]
    return value


# Definição dos arquivos de origem e destino
arquivo_distribuidor = sys.argv[2] 
arquivo_funcional = sys.argv[1] 
arquivo_pf_margem = sys.argv[3]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)

# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)

# Formatar CNPJ nas bases
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(remover_aspas_cnpj).apply(format_cnpj)
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(format_cnpj)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(remove_decimal_suffix).apply(remover_aspas_cnpj)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(format_cnpj).apply(remover_aspas_cnpj)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(remove_decimal_suffix).apply(remover_aspas_cnpj)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(format_cnpj).apply(remover_aspas_cnpj)

# Criar colunas específicas combinadas nas bases
print('Criando as colunas combinadas nas bases...')
base_funcional['CNPJ+NF'] = base_funcional['CNPJ'] + '-' + base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix)
base_funcional['CNPJ+NF+EAN'] = base_funcional['CNPJ+NF'] + '-' + base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) 
base_distribuidor['CNPJ+NF'] = base_distribuidor['CNPJ'] + '-' + base_distribuidor['NOTA'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['CNPJ+NF+EAN'] = base_distribuidor['CNPJ+NF'] + '-' + base_distribuidor['EAN'].astype(str).apply(remove_decimal_suffix)

    # Criar colunas específicas combinadas nas bases
print('Criando as colunas combinadas nas bases...')
base_funcional['CNPJ+NF'] = base_funcional['CNPJ'] + '-' + base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix)
base_funcional['CNPJ+NF+EAN'] = base_funcional['CNPJ+NF'] + '-' + base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['CNPJ+NF'] = base_distribuidor['CNPJ'] + '-' + base_distribuidor['NOTA'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['CNPJ+NF+EAN'] = base_distribuidor['CNPJ+NF'] + '-' + base_distribuidor['EAN'].astype(str).apply(remove_decimal_suffix)

# Criar as chaves e colunas adicionais nas bases
print('Criando as chaves nas bases...')

# Criar a quantidade somada antes de definir a chave para base funcional
base_funcional['Quantidade Somada'] = base_funcional.groupby(['CNPJ', 'Número da nota', 'EAN do produto'])['Quantidade Faturada'].transform('sum')
# Substituir NaN na coluna 'Quantidade Somada' com o valor de 'QTDE'
base_funcional['Quantidade Somada'] = base_funcional['Quantidade Somada'].fillna(base_funcional['Quantidade Faturada'])
# Criar a chave com a quantidade somada
base_funcional['Chave'] = (
    base_funcional['CNPJ'] + '-' +
    base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_funcional['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)

# Criar a quantidade somada antes de definir a chave para base distribuidor
base_distribuidor['Quantidade Somada'] = base_distribuidor.groupby(['CNPJ', 'NOTA', 'EAN'])['QTDE'].transform('sum')
# Substituir NaN na coluna 'Quantidade Somada' com o valor de 'QTDE'
base_distribuidor['Quantidade Somada'] = base_distribuidor['Quantidade Somada'].fillna(base_distribuidor['QTDE'])
# Criar a chave com a quantidade somada
base_distribuidor['Chave'] = (
    base_distribuidor['CNPJ'] + '-' +
    base_distribuidor['NOTA'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_distribuidor['EAN'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_distribuidor['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)

base_funcional['Status'] = ''
base_distribuidor['Status'] = ''

# Comparando as bases
print('Comparando as bases...')
base_distribuidor['Status'] = base_distribuidor['Chave'].isin(base_funcional['Chave']).map({True: 'OK', False: 'Chave não localizada'})
base_funcional['Status'] = base_funcional['Chave'].isin(base_distribuidor['Chave']).map({True: 'OK', False: 'Chave só consta na Funcional'})

# Lê a base pf_margem
pf_margem = pd.read_excel('pf_margem.xlsx', sheet_name='Dados')

# Formatar a coluna EAN na base pf_margem e na base_distribuidor
pf_margem['EAN'] = pf_margem['EAN'].astype(str).apply(remove_decimal_suffix)
base_distribuidor['EAN'] = base_distribuidor['EAN'].astype(str).apply(remove_decimal_suffix)

# Realiza o merge entre a base_distribuidor e a pf_margem usando o EAN
base_distribuidor = base_distribuidor.merge(
    pf_margem[['EAN', 'DESCONTO ENTRADA', 'MARGEM OL', 'PF']],
    left_on='EAN',
    right_on='EAN',
    how='left'
)

# Renomear as colunas conforme solicitado
base_distribuidor.rename(columns={
    'DESCONTO ENTRADA': 'Desconto Entrada MSD',
    'MARGEM OL': 'Margem MSD',
    'PF': 'PF MSD'
}, inplace=True)

# Continuar pegando o 'Desconto Funcional' da base_funcional
base_distribuidor = base_distribuidor.merge(
    base_funcional[['Chave', 'Desconto pedido']],
    on='Chave',
    how='left'
)

base_distribuidor.rename(columns={'Desconto pedido': 'Desconto Funcional'}, inplace=True)

#=(100-ICMS)%*(Desc. Pedido+MargemOL-Desc Entrada)%*Qnt Faturada* Preço Fábrica
# Adicionar a coluna 'Ressarcimento Funcional' com o cálculo fornecido usando a 'Qtd'
base_distribuidor['Ressarcimento Funcional'] = np.where(
    (100 - 18) / 100 * 
    (base_distribuidor['Desconto Funcional'] + base_distribuidor['Margem MSD'] - base_distribuidor['Desconto Entrada MSD']) / 100 * 
    base_distribuidor['QTDE'] *
    base_distribuidor['PF MSD'] < 0, 
    0,  # Se o valor for negativo, substitui por 0
    (100 - 18) / 100 * 
    (base_distribuidor['Desconto Funcional'] + base_distribuidor['Margem MSD'] - base_distribuidor['Desconto Entrada MSD']) / 100 * 
    base_distribuidor['QTDE'] *
    base_distribuidor['PF MSD']  # Caso contrário, mantém o valor calculado
)

# Reordenando colunas - Colocando as novas colunas no início
novas_colunas_funcional = [
    'Chave','Status'
]
outras_colunas_funcional = [col for col in base_funcional.columns if col not in novas_colunas_funcional]
base_funcional = base_funcional[novas_colunas_funcional + outras_colunas_funcional]

novas_colunas_distribuidor = [
    'Chave', 'Status'
]
outras_colunas_distribuidor = [col for col in base_distribuidor.columns if col not in novas_colunas_distribuidor]
base_distribuidor = base_distribuidor[novas_colunas_distribuidor + outras_colunas_distribuidor]

print('Exibindo uma simples amostra de como ficou...')
time.sleep(1)
print("\nBase Funcional:\n", base_funcional.head())
print("\nBase Distribuidor:\n", base_distribuidor.head())
print('Salvando...')

# Apagar as colunas CNPJ+NF, CNPJ+NF+EAN, e Quantidade Somada
base_distribuidor.drop(columns=['CNPJ+NF', 'CNPJ+NF+EAN', 'Quantidade Somada'], inplace=True)
# salvar em duas abas diferentes do mesmo arquivo

with pd.ExcelWriter('Conciliação Finalizada Agille.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print('Finalizado!')