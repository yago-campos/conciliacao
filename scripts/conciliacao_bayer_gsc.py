import pandas as pd
import openpyxl
import sys

# Funções de formatação e remoção de sufixo decimal
def format_CNPJ(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(value) if isinstance(value, str) else str(value)
    return value.zfill(14)

# Função para remover as aspas simples do CNPJ
def remover_aspas_CNPJ(df, coluna_CNPJ):
    # Remover aspas simples do CNPJ
    df[coluna_CNPJ] = df[coluna_CNPJ].astype(str).str.replace("'", "")
    return df

def remove_decimal_suffix(value):
    if isinstance(value, str) and value.endswith('.0'):
        return value[:-2]
    return value
    
# Definição dos arquivos de origem e destino
arquivo_funcional = sys.argv[1]
arquivo_distribuidor = sys.argv[2]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)

# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)

# Formatar CNPJ nas bases
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(format_CNPJ)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(format_CNPJ)

base_funcional['Quantidade Somada'] = base_funcional.groupby(['CNPJ', 'Número da nota', 'EAN do produto'])['Quantidade Faturada'].transform('sum')

# Criar as chaves e colunas adicionais nas bases
print('Criando as chaves nas bases...')
base_funcional['Chave'] = (
    base_funcional['CNPJ'] + '-' +
    base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_funcional['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)

base_distribuidor['Quantidade Somada'] = base_distribuidor.groupby(['CNPJ', 'Nr. Nota', 'Código EAN'])['Qtd'].transform('sum')

base_distribuidor['Chave'] = (
    base_distribuidor['CNPJ'] + '-' +
    base_distribuidor['Nr. Nota'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_distribuidor['Código EAN'].astype(str).apply(remove_decimal_suffix) + '-' +
    base_distribuidor['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)

base_funcional['Status'] = ''
base_distribuidor['Status'] = ''

# Comparando as chaves
print('Comparando as bases...')
base_distribuidor['Status'] = base_distribuidor['Chave'].isin(base_funcional['Chave']).map({True: 'OK', False: 'Chave não localizada'})
base_funcional['Status'] = base_funcional['Chave'].isin(base_distribuidor['Chave']).map({True: 'OK', False: 'Chave só consta na Funcional'})

'''
# Criar nova coluna 'Status Preço' para validar a diferença entre VPF e Preço Distribuidor na data do pedido
base_distribuidor['Status Preço'] = base_distribuidor.apply(
    lambda row: 'Preço diferente, abrir ticket para BD.' if row['VPF'] != base_funcional.loc[
        base_funcional['Chave'] == row['Chave'], 'Preço Distribuidor(data do pedido)'].values[0] else 'Preço OK',
    axis=1
)
'''

'''
# Realiza o merge entre base_distribuidor e base_funcional para trazer a coluna 'Monitorado'
base_distribuidor = base_distribuidor.merge(base_funcional[[
    'Chave', 'Monitorado']], on='Chave', how='left')

# Aplicar a fórmula para o cálculo do 'Ressarcimento Funcional' com base na condição de 'Monitorado'
base_distribuidor['Recálculo Distribuidor'] = base_distribuidor.apply(
    lambda row: (
        row['% Desc. Comercial Industria'] / 100 * row['VPF'] * row['Qtd']
        if row['Monitorado'] == 'Não' else
        row['% Desc. Comercial Industria'] / 100 * row['VPF'] * row['Qtd']
    ), axis=1
)

# Aplicar a lógica de cálculo de 'Ressarcimento Funcional' para a base_funcional
base_funcional['Ressarcimento Funcional'] = base_funcional.apply(
    lambda row: (
        (row['Menor desconto'] / 100) *
        row['Quantidade Faturada'] *
        (row['Preço Fábrica'] if row['Monitorado'] != 'Não' else row['VPF'])
    ), axis=1
)
'''
# Reordenando colunas - Colocando as novas colunas no início
novas_colunas_funcional = [
    'Chave','Status'
]
outras_colunas_funcional = [col for col in base_funcional.columns if col not in novas_colunas_funcional]
base_funcional = base_funcional[novas_colunas_funcional + outras_colunas_funcional]

novas_colunas_distribuidor = [
    'Chave', 'Status'
]
outras_colunas_distribuidor = [col for col in base_distribuidor.columns if col not in novas_colunas_distribuidor]
base_distribuidor = base_distribuidor[novas_colunas_distribuidor + outras_colunas_distribuidor]

print('Salvando...')

# Apagar as colunas CNPJ+NF, CNPJ+NF+EAN, e Quantidade Somada
base_distribuidor.drop(columns=['CNPJ+NF', 'CNPJ+NF+EAN', 'Quantidade Somada'], inplace=True)

# salvar em duas abas diferentes do mesmo arquivo
with pd.ExcelWriter('Conciliação Finalizada Santa Cruz.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print('Finalizado!')