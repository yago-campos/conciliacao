import pandas as pd
import openpyxl
import os
import time
import sys

# Função para formatar o CNPJ para 14 dígitos
def format_cnpj(value):
    if pd.isna(value):
        return '00000000000000'
    value = str(int(value)) if isinstance(value, (int, float)) else str(value)
    return value.zfill(14)

# Função para remover o sufixo ".0"
def remove_decimal_suffix(value):
    if isinstance(value, str) and value.endswith('.0'):
        return value[:-2]
    return value

# Arquivos Excel
arquivo_distribuidor = sys.argv[2]
arquivo_funcional = sys.argv[1]

# 1. Carregar o arquivo baseDistribuidor
print('Lendo o arquivo do distribuidor...')
try:
    base_distribuidor = pd.read_excel(arquivo_distribuidor, skiprows=1)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseDistribuidor: {e}")
    sys.exit(1)

# 2. Carregar o arquivo baseFuncional usando OpenPyxl
print('Lendo o arquivo baseFuncional...')
try:
    arquivo = openpyxl.load_workbook(arquivo_funcional)
    print('Selecionando a aba ativa...')
    planilha = arquivo.active
    tabela_intervalo = f'A1:{planilha.cell(row=1, column=planilha.max_column).column_letter}{planilha.max_row}'
    tabela = openpyxl.worksheet.table.Table(displayName="TabelaSimples", ref=tabela_intervalo)
    planilha.add_table(tabela)
    dados = planilha.values
    colunas = next(dados)  # A primeira linha é o cabeçalho
    base_funcional = pd.DataFrame(dados, columns=colunas)
except Exception as e:
    print(f"Erro ao carregar o arquivo baseFuncional: {e}")
    sys.exit(1)

# Formatando CNPJ na base Funcional e Distribuidor para 14 dígitos
base_funcional['CNPJ'] = base_funcional['CNPJ'].apply(format_cnpj)
base_distribuidor['CNPJ'] = base_distribuidor['CNPJ'].apply(format_cnpj)

print('Removendo pontos, traços e barras da base do distribuidor...')

# Criando as chaves na base Funcional
print('Criando as chaves na base Funcional...')
base_funcional['Chave1'] = (
base_funcional['CNPJ'] + '-' +
base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['Quantidade Faturada'].astype(str).apply(remove_decimal_suffix)
)
base_funcional['Quantidade Somada'] = base_funcional.groupby(['CNPJ', 'Número da nota', 'EAN do produto'])['Quantidade Faturada'].transform('sum')
base_funcional['Chave2'] = (
base_funcional['CNPJ'] + '-' +
base_funcional['Número da nota'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['EAN do produto'].astype(str).apply(remove_decimal_suffix) + '-' +
base_funcional['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)
base_funcional['Status'] = ''
base_funcional['Status1'] = '' # excluir no final
base_funcional['Status2'] = '' # excluir no final
base_funcional['Validação V/F'] = '' # excluir no final
base_funcional['AÇÃO'] = ''
base_funcional['ORIGEM'] = ''

# Criando as chaves na base do distribuidor
print('Criando as chaves na base do distribuidor...')
base_distribuidor['Chave1'] = (
base_distribuidor['CNPJ'] + '-' +
base_distribuidor['NFe'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['EAN mer'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['Quantidade'].astype(str).apply(remove_decimal_suffix)
)
base_distribuidor['Quantidade Somada'] = base_distribuidor.groupby(['CNPJ', 'NFe', 'EAN mer'])['Quantidade'].transform('sum')
base_distribuidor['Chave2'] = (
base_distribuidor['CNPJ'] + '-' +
base_distribuidor['NFe'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['EAN mer'].astype(str).apply(remove_decimal_suffix) + '-' +
base_distribuidor['Quantidade Somada'].astype(str).apply(remove_decimal_suffix)
)
base_distribuidor['Status'] = ''
base_distribuidor['Status1'] = '' # excluir no final
base_distribuidor['Status2'] = '' # excluir no final
base_distribuidor['Validação V/F'] = '' # excluir no final
base_distribuidor['AÇÃO'] = ''
base_distribuidor['ORIGEM'] = ''

print('Comparando as bases...')

base_distribuidor['Status1'] = base_distribuidor['Chave1'].isin(base_funcional['Chave1']).map({True: 'OK', False: 'NF não localizada'})
base_funcional['Status1'] = base_funcional['Chave1'].isin(base_distribuidor['Chave1']).map({True: 'OK', False: 'NF só consta na Funcional'})

base_distribuidor['Status2'] = base_distribuidor['Chave2'].isin(base_funcional['Chave2']).map({True: 'OK', False: 'NF não localizada'})
base_funcional['Status2'] = base_funcional['Chave2'].isin(base_distribuidor['Chave2']).map({True: 'OK', False: 'NF só consta na Funcional'})

base_funcional['Validação V/F'] = base_funcional['Status1'].eq(base_funcional['Status2'])
base_distribuidor['Validação V/F'] = base_distribuidor['Status1'].eq(base_distribuidor['Status2'])

print('Salvando...')

# salvar em duas abas diferentes do mesmo arquivo
with pd.ExcelWriter('Conciliacao_Finalizada_Santa_Cruz.xlsx', engine='openpyxl') as writer:
    base_distribuidor.to_excel(writer, sheet_name='Distribuidor', index=False)
    base_funcional.to_excel(writer, sheet_name='Funcional', index=False)

print('Finalizado!')